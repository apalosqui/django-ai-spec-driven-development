from __future__ import annotations

from typing import Any, Dict, List, Optional, Union


def _select_sheet(wb, sheet: Union[str, int, None]):
    """
    Pick a worksheet by name, index, or the first sheet if None.
    """
    if sheet is None:
        return wb.worksheets[0]
    if isinstance(sheet, int):
        try:
            return wb.worksheets[sheet]
        except IndexError as exc:
            raise ValueError(f'sheet index out of range: {sheet}') from exc
    try:
        return wb[sheet]
    except KeyError as exc:
        raise ValueError(f'sheet not found: {sheet!r}') from exc


def read_sheet_as_dicts(
    path: str,
    sheet: Union[str, int, None] = None,
    *,
    header: bool = True,
    skip_blank_rows: bool = True,
    data_only: bool = True,
    read_only: bool = True,
) -> List[Dict[str, Any]]:
    """
    Read an Excel sheet into a list of dicts.

    Notes about formulas:
    - openpyxl does not calculate formulas; with `data_only=True`, it returns
      the cached result saved in the file by Excel/LibreOffice. If the file
      was never opened/saved by a spreadsheet app, cached values may be None.

    Parameters
    - path: File path to the workbook.
    - sheet: Sheet name, zero-based index, or None for the first sheet.
    - header: If True, first row is used as column names. If a header cell is
      empty, a fallback like 'col_1' is used.
    - skip_blank_rows: Skip rows where all values are empty/None.
    - data_only: If True, return cached results for formulas.
    - read_only: Use openpyxl read-only mode to reduce memory.

    Returns
    - List of dictionaries: one per row.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - import guard
        raise RuntimeError(
            'openpyxl is required. Install with '\
            '`pip install -r requirements.txt -r requirements-excel.txt`'
        ) from exc

    wb = load_workbook(filename=path, data_only=data_only, read_only=read_only)
    try:
        ws = _select_sheet(wb, sheet)
        rows = ws.iter_rows(values_only=True)

        headers: List[str]
        items: List[Dict[str, Any]] = []

        if header:
            try:
                raw = next(rows)
            except StopIteration:
                return []
            headers = [
                (str(val).strip() if (val is not None and str(val).strip()) else f'col_{idx + 1}')
                for idx, val in enumerate(raw)
            ]
        else:
            # If no header, generate generic headers based on first row width
            try:
                first = next(rows)
            except StopIteration:
                return []
            headers = [f'col_{i + 1}' for i in range(len(first))]
            # process first row as data
            if not (skip_blank_rows and _is_blank_row(first)):
                items.append({headers[i]: first[i] for i in range(len(headers))})

        for row in rows:
            if skip_blank_rows and _is_blank_row(row):
                continue
            items.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})

        return items
    finally:
        wb.close()


def _is_blank_row(row: tuple) -> bool:
    return all(_is_empty_cell(v) for v in row)


def _is_empty_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False

